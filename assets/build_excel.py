#!/usr/bin/env python3
"""
build_excel.py — generates the pricing-discipline merged workbook.
One sheet per block (Price, Growth Quality, Capital, Defensibility, Decision) + Summary.
Cell-color convention (Abdulrahman's standard): blue inputs, black formulas, yellow outputs.
Source flags as cell comments: public-comp URL / "MGMT" / "ESTIMATE".

Usage:
  python build_excel.py --config deal.json --out pricing-{slug}.xlsx
deal.json carries every block input the HTML tool collected (see references/excel-structure.md).
The math here MUST match references/frameworks.md and the HTML engine.
"""
import json, argparse, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
try:
    from validate_deal import check as _validate
except Exception:
    _validate = None

# ---- shared engine (mirror of frameworks.md + sourced-multiples.md) ----
# Sourced multiples come from cfg["tiers"]: list of {name,min,mult,exitMult}.
# No fabricated fallback ladder — tiers are required and must be sourced (the model's #1 rule).
def _tiers(cfg):
    t=cfg.get("tiers")
    if not t:
        raise SystemExit("config error: 'tiers' is required (sourced sector medians) — "
                         "no fabricated fallback. See references/sourced-multiples.md.")
    # defensively sort high->low by min so tier_for is order-independent
    return sorted(t, key=lambda x: x.get("min",0), reverse=True)
def tier_for(cfg,g):
    for t in _tiers(cfg):
        if g>=t["min"]: return t
    return _tiers(cfg)[-1]
def tier_name(cfg,g): return tier_for(cfg,g)["name"]
def entry_mult(cfg,g): return tier_for(cfg,g)["mult"]
def exit_multiple(cfg,g):
    # mirror the HTML engine: exitMult if explicitly set (non-null), else fall back to mult
    # (which may itself be None -> sector valued on EV/EBITDA, "FCF" exit).
    t=tier_for(cfg,g); em=t.get("exitMult")
    return em if em is not None else t.get("mult")
def exit_tier_name(cfg,g): return tier_for(cfg,g)["name"]

# Default registry — MUST mirror the HTML template's default SOURCES so the Excel "Sources" sheet and
# the HTML "Benchmarking" tab cite the same registry for the same deal. A deal can override both by
# carrying its own cfg["sources"] (same shape as the HTML: {n,label,url,note,kind:"index"|"convention"}).
SOURCES=[
 (1,"Finro — Fintech Valuation Multiples Q1 2026","https://www.finrofca.com/news/fintech-valuation-multiples-q1-2026","416-company fintech dataset. Blockchain/digital-asset median 14.2x EV/Revenue (avg 26.6x — use median). Payments 3.6x, WealthTech 16.2x, Lending ~2.5x.","live"),
 (2,"valueaddvc — Public SaaS Valuations","https://valueaddvc.com/saas-valuations","Public SaaS median ~8.5x EV/NTM (mid-2026). >40% growth = 2-3x slow-growth multiple; NRR <100% cuts 30-50%. AI-native 15-40x.","live"),
 (3,"multiples.vc — Coinbase comp","https://multiples.vc/public-comps/coinbase-valuation-multiples","Listed digital-asset infra reference: ~$47B EV, 75% GM / 39% EBITDA, revenue -15% YoY (May 2026). Exit multiples compress as growth decays.","live"),
 (4,"Aswath Damodaran — NYU Stern data","https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datacurrent.html","Industry EV/EBITDA, EV/Sales, margins (Jan 2026). Marketplace, healthcare, hardware rocks.","live"),
 (5,"Windsor Drake — Fintech & SaaS Multiples 2026","https://windsordrake.com/fintech-valuation-multiples/","Payments 4-6x revenue / 8-12x EBITDA for established operators; lending ~2.5x. Private/public gap ~30-50%.","live"),
 (6,"David Sacks — 'The Burn Multiple' (Craft Ventures, 2020)","https://medium.com/craft-ventures/the-burn-multiple-51a7e43cb200","Burn multiple bands: <1x amazing / 1-2x good / 2-3x suspect / >3x dangerous. GM-adjust below 75%.","live"),
 (7,"Lemkin / O'Driscoll pricing convention","","85% decay, 1-in-3 reaccel, 1% TAM, integration thresholds, Fortnite, convergence, return-on-regulatory-capital. Growth-tier boundaries are convention; the multiples attached are sourced (1-5). Judgment, not data.","convention"),
]
def _sources(cfg):
    """Deal-supplied registry (HTML-shaped dicts) if present, else the default — normalized to
    (n, label, url, note, is_live) tuples for the Sources sheet."""
    s=cfg.get("sources")
    if isinstance(s,list) and s:
        return [(d.get("n"), d.get("label",""), d.get("url"), d.get("note",""), d.get("kind","index")!="convention") for d in s]
    return [(n,l,u,note,kind=="live") for (n,l,u,note,kind) in SOURCES]

# ---- styles ----
BLUE   = Font(color="1155CC", bold=False)            # input
BLACK  = Font(color="000000")                        # formula/derived
YELLOWF= PatternFill("solid", fgColor="FFF2CC")      # output highlight
HEADF  = PatternFill("solid", fgColor="0F1719")
HEAD   = Font(color="FFFFFF", bold=True, size=12)
SUB    = Font(color="6B7682", italic=True, size=9)
BOLD   = Font(bold=True)
TITLE  = Font(bold=True, size=14, color="0B6E6E")
thin   = Side(style="thin", color="E4E8ED")
BORD   = Border(bottom=thin)
RIGHT  = Alignment(horizontal="right")

def sheet_header(ws, title, sub):
    ws.merge_cells("A1:F1"); c=ws["A1"]; c.value=title; c.font=HEAD; c.fill=HEADF
    ws.merge_cells("A2:F2"); ws["A2"].value=sub; ws["A2"].font=SUB
    ws.column_dimensions["A"].width=34
    for col in "BCDEF": ws.column_dimensions[col].width=16

def put(ws, row, label, value, kind="formula", flag=None, fmt=None):
    ws.cell(row=row, column=1, value=label).font=BLACK
    c=ws.cell(row=row, column=2, value=value)
    c.font = BLUE if kind=="input" else BLACK
    if kind=="output": c.fill=YELLOWF; c.font=BOLD
    if fmt: c.number_format=fmt
    c.alignment=RIGHT
    if flag: c.comment=Comment(flag, "pricing-discipline")
    return row+1

def build(cfg):
    wb=Workbook()
    g0   = cfg["growth"]
    # real growth drives the cascade. Prefer the Block-2 waterfall (same as the HTML engine),
    # then an explicit realGrowth override, then headline growth as the last resort.
    _wf  = cfg.get("waterfall") or {}
    _ws  = _wf.get("start",0) or 0
    if _ws>0:
        real = (_wf.get("new",0)+_wf.get("exp",0)-_wf.get("churn",0)-_wf.get("down",0))/_ws
    else:
        real = cfg.get("realGrowth", g0)
    # multiples now sourced per tier from cfg["tiers"] (bedrock/band/scalar removed)
    R0   = cfg["revenue"]; P=cfg["askedPost"]; T=cfg["target"]
    d    = cfg.get("decay",0.85); H=cfg.get("hold",7)
    src  = cfg.get("tierSource") or cfg.get("bedrockSource","sourced median — see sourced-multiples.md")
    # single source of truth for ownership (mirrors the HTML: diluted = entry × (1 − dilution))
    own_   = cfg.get("entryOwnership",0.12); dil_pct=cfg.get("dilution",0.5)
    dOwn   = own_*(1-dil_pct); check=cfg.get("checkSize",3)

    # ---------- SUMMARY ----------
    ws=wb.active; ws.title="Summary"
    ws["A1"].value=f"Pricing Discipline — {cfg['name']}"; ws["A1"].font=TITLE
    ws["A2"].value=f"{cfg.get('assetClass','')} · {'Mark-to-reality' if cfg.get('mode')=='mtr' else 'Entry pricing'} · sector: {cfg.get('sectorLabel','')}"; ws["A2"].font=SUB
    ws.column_dimensions["A"].width=30
    for col in "BCDEF": ws.column_dimensions[col].width=16
    # recompute the cascade
    gEntry=real
    Me=entry_mult(cfg,gEntry); Ma=P/R0 if R0 else 0; rich=(Ma/Me-1) if Me else 0
    rev=R0
    for y in range(1,H+1): rev*=1+gEntry*(d**y)
    gH=gEntry*(d**H); Mx=exit_multiple(cfg,gH); fcf=(Mx is None)
    Vx=(Mx or 0)*rev
    # diluted-ownership MOIC = the post-cascade final state, matching the HTML after Block 3
    moic=(Vx*dOwn)/check if check else 0; ownNote="diluted"
    if fcf: verdict="Doesn't clear — exit decays into FCF"
    elif moic>=T: verdict="Clears"
    elif moic>=0.6*T: verdict="Tight"
    else: verdict="Doesn't clear — pass on price"
    r=4
    r=put(ws,r,"Asked multiple (x)", round(Ma,1), "formula", fmt="0.0")
    r=put(ws,r,"Sector median (entry tier, x)", round(Me,1) if Me else "EBITDA", "input", src, fmt="0.0")
    r=put(ws,r,"Entry tier", tier_name(cfg,gEntry), "formula", "growth tier (sourced)")
    r=put(ws,r,"Rich/cheap vs grid", round(rich,3), "output", fmt="0.0%")
    r=put(ws,r,"Headline growth", g0, "input","MGMT", fmt="0%")
    r=put(ws,r,"Real growth (Block 2)", real, "formula","price-stripped", fmt="0%")
    r=put(ws,r,"Decay rate", d, "input","Block 4 erosion read", fmt="0%")
    r=put(ws,r,f"Exit growth (Y{H})", round(gH,3), "formula", fmt="0%")
    r=put(ws,r,"Exit multiple (x)", "EBITDA" if fcf else round(Mx,1), "formula","sourced median, exit tier", fmt="0.0")
    r=put(ws,r,"Exit value ($M)", round(Vx,0), "formula", fmt="#,##0")
    r=put(ws,r,"Diluted ownership", round(dOwn,4), "formula", "entry × (1 − dilution)", fmt="0.0%")
    r=put(ws,r,f"MOIC ({ownNote})", round(moic,2), "output", fmt="0.00")
    r=put(ws,r,"Target MOIC", T, "input", fmt="0.0")
    ws.cell(row=r,column=1,value="Verdict").font=BOLD
    vc=ws.cell(row=r,column=2,value=verdict); vc.font=BOLD; vc.fill=YELLOWF

    # ---------- BLOCK 1 PRICE ----------
    ws=wb.create_sheet("1·Price")
    sheet_header(ws,"Block 1 — Price","Sourced sector median → tier → entry → decay → exit → MOIC")
    r=4
    r=put(ws,r,"Sector",cfg.get("sectorLabel",""),"input",src)
    r=put(ws,r,"Entry tier median (x)",round(Me,1) if Me else "EBITDA","output",src,fmt="0.0")
    r=put(ws,r,"Revenue ($M)",R0,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Asked post ($M)",P,"input","term sheet",fmt="#,##0")
    r=put(ws,r,"Asked multiple (x)",round(Ma,1),"formula",fmt="0.0")
    r=put(ws,r,"Entry tier",tier_name(cfg,gEntry),"formula")
    r=put(ws,r,"Entry multiple (x)",round(Me,1) if Me else "EBITDA","output",fmt="0.0")
    r=put(ws,r,"Rich/cheap vs grid",round(rich,3),"output",fmt="0.0%")
    r+=1
    ws.cell(row=r,column=1,value="Decay walk").font=BOLD; r+=1
    for h,lab in [(1,"Year"),(2,"Growth"),(3,"Exit tier"),(4,"Exit mult"),(5,"Revenue $M"),(6,"Value $M")]:
        cc=ws.cell(row=r,column=h,value=lab); cc.font=BOLD; cc.border=BORD
    r+=1
    rr=R0
    for y in range(0,H+1):
        gy=gEntry*(d**y) if y>0 else gEntry
        if y>0: rr*=1+gy
        em=exit_multiple(cfg,gy)
        ws.cell(row=r,column=1,value=f"Y{y}")
        ws.cell(row=r,column=2,value=round(gy,3)).number_format="0%"
        ws.cell(row=r,column=3,value=exit_tier_name(cfg,gy))
        ws.cell(row=r,column=4,value=("FCF" if em is None else round(em,1)))
        ws.cell(row=r,column=5,value=round(rr,0)).number_format="#,##0"
        ws.cell(row=r,column=6,value=("" if em is None else round(em*rr,0))).number_format="#,##0"
        r+=1
    r=put(ws,r,f"MOIC ({ownNote})",round(moic,2),"output",fmt="0.00")

    # ---------- BLOCK 2 GROWTH QUALITY ----------
    ws=wb.create_sheet("2·Growth Quality")
    sheet_header(ws,"Block 2 — Growth Quality","Real vs fake growth, maiming, reacceleration, TAM")
    w=cfg.get("waterfall",{}); s=w.get("start",R0)
    nl=w.get("new",0); ex=w.get("exp",0); ch=w.get("churn",0); dn=w.get("down",0); pr=w.get("price",0)
    realg=(nl+ex-ch-dn)/s if s else 0; fakeg=pr/s if s else 0
    grr=(s-ch-dn)/s if s else 0; nrr=(s-ch-dn+ex)/s if s else 0
    r=4
    r=put(ws,r,"Starting ARR ($M)",s,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"New logo ($M)",nl,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Expansion ($M)",ex,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Churn ($M)",ch,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Downgrades ($M)",dn,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Price increases ($M)",pr,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Real (organic) growth",round(realg,3),"output","excludes price",fmt="0%")
    r=put(ws,r,"Fake (price) growth",round(fakeg,3),"formula",fmt="0%")
    r=put(ws,r,"GRR",round(grr,3),"formula",fmt="0%")
    r=put(ws,r,"NRR",round(nrr,3),"formula",fmt="0%")
    r=put(ws,r,"Real-growth tier",tier_name(cfg,realg),"output")
    n0=cfg.get("nrr0"); n1=cfg.get("nrr1")
    if n0 and n1:
        r=put(ws,r,"NRR 2y ago",n0/100,"input","MGMT",fmt="0%")
        r=put(ws,r,"NRR today",n1/100,"input","MGMT",fmt="0%")
        maim = (grr>=0.90 and (n1-n0)<=-8)
        r=put(ws,r,"Maiming signature",("YES — push decay 70-75%" if maim else "no"),"output")

    # ---------- BLOCK 3 CAPITAL ----------
    ws=wb.create_sheet("3·Capital")
    sheet_header(ws,"Block 3 — Capital","Burn multiple (GM-adjusted) + fund math at diluted ownership")
    burn=cfg.get("burn",0); nn=cfg.get("netNewArr",0); gm=cfg.get("gm",0.75)
    raw=burn/nn if nn else 0; adj=burn/(nn*gm) if (nn and gm) else 0
    own=own_; dilpct=dil_pct; fund=cfg.get("fundSize",150)   # dOwn already computed once at top of build()
    r=4
    r=put(ws,r,"Net burn 12M ($M)",burn,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Net new ARR 12M ($M)",nn,"input","MGMT",fmt="#,##0.0")
    r=put(ws,r,"Gross margin",gm,"input","MGMT",fmt="0%")
    r=put(ws,r,"Raw burn multiple (x)",round(raw,2),"formula",fmt="0.00")
    r=put(ws,r,"GM-adjusted burn (x)",round(adj,2),"output","÷(net new × GM)",fmt="0.00")
    r+=1
    r=put(ws,r,"Entry ownership",own,"input","term sheet",fmt="0.0%")
    r=put(ws,r,"Future dilution",dilpct,"input","ESTIMATE",fmt="0%")
    r=put(ws,r,"Diluted ownership",round(dOwn,4),"output",fmt="0.00%")
    r=put(ws,r,"Check size ($M)",check,"input",fmt="#,##0.0")
    r=put(ws,r,"Fund size ($M)",fund,"input",fmt="#,##0")
    r+=1
    ws.cell(row=r,column=1,value="Fund-returner scenarios").font=BOLD; r+=1
    for h,lab in [(1,"Scenario"),(2,"Exit ARR"),(3,"Mult"),(4,"Exit $M"),(5,"Proceeds"),(6,"MOIC/check")]:
        cc=ws.cell(row=r,column=h,value=lab); cc.font=BOLD; cc.border=BORD
    r+=1
    for nm,arr,m in [("Grind",R0*4,5),("Solid",R0*10,8),("Breakout",R0*20,12)]:
        ev=arr*m; proc=ev*dOwn; mo=proc/check if check else 0
        ws.cell(row=r,column=1,value=nm); ws.cell(row=r,column=2,value=round(arr,0)).number_format="#,##0"
        ws.cell(row=r,column=3,value=f"{m}x"); ws.cell(row=r,column=4,value=round(ev,0)).number_format="#,##0"
        ws.cell(row=r,column=5,value=round(proc,1)).number_format="#,##0.0"; ws.cell(row=r,column=6,value=round(mo,1)).number_format='0.0"x"'
        r+=1

    # ---------- BLOCK 4 DEFENSIBILITY ----------
    ws=wb.create_sheet("4·Defensibility")
    sheet_header(ws,"Block 4 — Defensibility","Four-dimension moat score + erosion + convergence")
    moat=(list(cfg.get("moat") or [])+[3,3,3,3])[:4]   # pad/truncate to exactly 4 so a short array can't IndexError
    r=4
    for i,dim in enumerate(["Technology","Distribution","Network effects","Switching costs"]):
        r=put(ws,r,dim,moat[i],"input","1-5 score",fmt="0")
    r=put(ws,r,"Moat average",round(sum(moat)/4,1),"output",fmt="0.0")
    r=put(ws,r,"Erosion rate",cfg.get("erosion","moderate"),"input")
    r=put(ws,r,"Convergence layer",cfg.get("convergenceLayer","point solution"),"input")
    ig=cfg.get("integAvg",0); igp=cfg.get("integPct",0)
    r=put(ws,r,"Integrations/customer",ig,"input","MGMT",fmt="0")
    r=put(ws,r,"% customers using 3+",igp/100 if igp else 0,"input","MGMT",fmt="0%")

    # ---------- BLOCK 5 DECISION ----------
    ws=wb.create_sheet("5·Decision")
    sheet_header(ws,"Block 5 — Decision","Three-Box gate, entry-vs-TAM, recommended number")
    r=4
    tb=cfg.get("threeBox",{})
    r=put(ws,r,"Box 1 — Market",("pass" if tb.get("market") else "OPEN"),"input")
    r=put(ws,r,"Box 2 — Team",("pass" if tb.get("team") else "OPEN"),"input")
    r=put(ws,r,"Box 3 — Residual risk",("pass" if tb.get("risk") else "OPEN"),"input")
    tam=cfg.get("tam",0)
    if tam: r=put(ws,r,"Entry as % of TAM",round(P/tam,3),"formula","entry≥TAM ⇒ unprofitable",fmt="0%")
    recPost = 0 if fcf else (Vx*dOwn/T)
    r=put(ws,r,"Recommended post ($M)",("—" if fcf else round(recPost,0)),"output",fmt="#,##0")
    r=put(ws,r,"Recommended multiple (x)",("—" if (fcf or not R0) else round(recPost/R0,1)),"formula",fmt="0.0")
    ws.cell(row=r,column=1,value="Verdict").font=BOLD
    vc=ws.cell(row=r,column=2,value=verdict); vc.font=BOLD; vc.fill=YELLOWF

    # ---------- SOURCES ----------
    ws=wb.create_sheet("Sources")
    sheet_header(ws,"Benchmarking & Sources","Every benchmark traced; live indices linked, conventions marked")
    ws.column_dimensions["A"].width=6; ws.column_dimensions["B"].width=42
    ws.column_dimensions["C"].width=52; ws.column_dimensions["D"].width=12
    r=4
    for h,lab in [(1,"#"),(2,"Source"),(3,"Anchors"),(4,"Type")]:
        cc=ws.cell(row=r,column=h,value=lab); cc.font=BOLD; cc.border=BORD
    r+=1
    for n,label,url,note,is_live in _sources(cfg):
        ws.cell(row=r,column=1,value=n).font=BLACK
        lc=ws.cell(row=r,column=2,value=label)
        if url: lc.hyperlink=url; lc.font=Font(color="0E8A8A",underline="single")
        ws.cell(row=r,column=3,value=note).font=BLACK
        kc=ws.cell(row=r,column=4,value=("live source" if is_live else "convention"))
        kc.font=Font(color=("0B6E6E" if is_live else "B4540A"),bold=True)
        r+=1
    return wb

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--config",required=True)
    ap.add_argument("--out",required=True)
    a=ap.parse_args()
    cfg=json.load(open(a.config, encoding="utf-8"))
    if _validate:
        _, warns = _validate(cfg)
        for w in warns: print("warning:", w, file=sys.stderr)
    wb=build(cfg)
    wb.save(a.out)
    print("wrote",a.out)

if __name__=="__main__":
    main()
